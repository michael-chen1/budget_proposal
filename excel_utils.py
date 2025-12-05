from openpyxl import load_workbook
from numbers import Number

def coerce_excel_value(v):
    """
    Convert numeric-looking strings to int/float; leave everything else as-is.
    """
    # Already numeric → keep it
    if isinstance(v, Number):
        return v

    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None  # empty string → blank cell

        # Remove common formatting like commas
        s_no_commas = s.replace(",", "")

        # Try int/float parse
        try:
            if "." in s_no_commas:
                return float(s_no_commas)
            else:
                return int(s_no_commas)
        except ValueError:
            # Not a pure number → leave as text
            return v

    # Other types (None, bool, dates, etc.) just pass through
    return v


def populate_template(extracted: dict, template_path: str, output_path: str):
    """
    Open the Excel file at `template_path`, and for each key in `extracted`,
    write its value into the single cell defined by the named range of the same name.
    """
    wb = load_workbook(template_path)

    for key, value in extracted.items():
        # Skip keys that aren't defined names in the workbook
        if key not in wb.defined_names:
            continue

        dn = wb.defined_names[key]
        coerced = coerce_excel_value(value)

        for sheet_name, coord in dn.destinations:
            ws = wb[sheet_name]
            ws[coord] = coerced  # now a real number if it looked numeric

    wb.save(output_path)
