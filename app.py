from dotenv import load_dotenv
load_dotenv()

import os
from flask import Flask, request, render_template, session, send_file, redirect, url_for, jsonify
from extractors import get_data_biostats, calculate_dmc, calculate_refresh, get_data_dm, get_data_pm, get_data_conform
from excel_utils import populate_template
from openpyxl import Workbook, load_workbook
import tempfile
import json


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_KEY", "edetek123")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

ALLOWED_EXTENSIONS = {"pdf", "docx"}


# Short descriptions for each extracted field. Any field not listed here will
# simply render with a blank description cell in the results table.
FIELD_DESCRIPTIONS = {
    "adam_compl": "Number of complex ADaM datasets",
    "adam_dmc_fr": "Number of full refreshes for ADaM datasets relating to DMC meetings",
    "adam_ia_fr": "Number of full refreshes for ADaM datasets relating to IA",
    "adam_fr": "Number of full refreshes for ADaM datasets",
    "adam_simp": "Number of simple ADaM datasets",
    "anaylsis_dur": "Duration of analysis phase of study (months)",
    "auto_queries_complete": "Number of auto queries per completed subject",
    "auto_queries_screen_fail": "Number of auto queries per screen failure subject",
    "auto_queries_total": "Total number of auto queries performed",
    "auto_queries_withdrawn": "Number of auto queries per withdrawn subject",
    "close_dur": "Duration of close-out phase of study (months) ",
    "crf_pages_per_visit": "Number of CRF pages per visit",
    "crf_pages_complete": "Number of CRF pages per completed subject",
    "crf_pages_screen_fail": "Number of CRF pages per screen failure subject",
    "crf_pages_total": "Total number of CRF pages",
    "crf_pages_withdrawn": "Number of CRF pages per withdrawn subject",
    "data_review_listings": "Total number of manual data review listings",
    "dmc/ia": "Boolean value for if DMC meetings or Interim Analysis (IA) is required",
    "dropout_rate": "Dropout rate of subjects",
    "dsur_datasets": "Number of datasets needed for the Development Safety Update Report (DSUR)",
    "dsur_report_listings": "Number of listings needed for the Development Safety Update Report (DSUR)",
    "dsur_report_tables": "Number of tables needed for the Development Safety Update Report (DSUR)",
    "dsur_years": "Number of annual refreshes needed for the Development Safety Update Report (DSUR)",
    "enroll_dur": "Duration of enrollment phase of study (months)",
    "external_data_reconcilation": "Number of instances of external data reconcilation needed",
    "investigator_datasets": "Number of datasets needed for the Investigator's Brochure (IB)",
    "investigator_listings": "Number of listings needed for the Investigator's Brochure (IB)",
    "investigator_tables": "Number of tables needed for the Investigator's Brochure (IB)",
    "investigator_years": "Number of annual refreshes needed for the Investigator's Brochure (IB)",
    "manual_queries_complete": "Number of manual queries per completed subject",
    "manual_queries_total": "Total number of manual queries performed",
    "manual_queries_withdrawn": "Number of manual queries per withdrawn subject",
    "num_complete": "Number of completed subjects",
    "num_countries": "Number of countries",
    "num_data_metrics_report": "Number of data metrics reports",
    "num_dmc_meet": "Number of DMC meetings",
    "num_external_data_source": "Number of external data sources",
    "num_lab_panel": "Number of lab panels for each local lab",
    "num_local_lab": "Number of local labs",
    "num_meetings": "",
    "num_sae": "Number of serious adverse events",
    "num_screen_fail": "Number of screen failure subjects",
    "num_screened_subj": "Number of subjects screened",
    "num_sites": "Number of sites",
    "num_subj": "Number of enrolled subjects",
    "num_unique_terms_aemh": "Number of unique AE and MH terms",
    "num_unique_terms_cm": "Number of unique CM terms",
    "num_withdrawn": "Number of withdrawn subjects",
    "patient_profile": "Number of patient profiles needed",
    "prog_support_requests": "Number of programming support hours needed",
    "protocol_deviation_check": "",
    "safety_signal_report": "Number of quarterly reports for safety signal detection",
    "screen_failure_rate": "Screen failure rate (%)",
    "sdtm_dmc_fr": "Number of full refreshes for SDTM datasets relating to DMC meetings",
    "sdtm_ia_fr": "Number of full refreshes for SDTM datasets relating to IA",
    "sdtm_fr": "Number of full refreshes for SDTM datasets",
    "sdtm_sd": "Number of SDTM subject domains",
    "sdtm_tdd": "Number of SDTM trial design domains",
    "start_dur": "Duration of start-up phase of study (months)",
    "stat_support_requests": "Number of statistician support hours needed",
    "subj_dur": "Duration of subject participation phase of study (months)",
    "tlf_dmc_fr": "Number of full refreshes for TLFs relating to DMC meetings",
    "tlf_dmc_repeat_figures": "Number of repeat figures needed for DMC meetings",
    "tlf_dmc_repeat_listings": "Number of repeat listings needed for DMC meetings",
    "tlf_dmc_repeat_tables": "Number of repeat tables needed for DMC meetings",
    "tlf_dmc_unique_figures": "Number of unique figures needed for DMC meetings",
    "tlf_dmc_unique_listings": "Number of unique listings needed for DMC meetings",
    "tlf_dmc_unique_tables": "Number of unique tables needed for DMC meetings",
    "tlf_final_fr": "Number of full refreshes for TLFs",
    "tlf_final_repeat_figures": "Number of repeat figures needed for study",
    "tlf_final_repeat_listings": "Number of repeat listings needed for study",
    "tlf_final_repeat_tables": "Number of repeat tables needed for study",
    "tlf_final unique_figures": "Number of unique figures needed for study",
    "tlf_final_unique_listings": "Number of unique listings needed for study",
    "tlf_final_unique_tables": "Number of unique tables needed for study",
    "tlf_ia_fr": "Number of full refreshes for TLFs relating to Interim Analysis (IA)",
    "tlf_ia_repeat_figures": "Number of repeat figures needed for Interim Analysis (IA)",
    "tlf_ia_repeat_listings": "Number of repeat listings needed for Interim Analysis (IA)",
    "tlf_ia_repeat_tables": "Number of repeat tables needed for Interim Analysis (IA)",
    "tlf_ia_unique_figures": "Number of unique figures needed for Interim Analysis (IA)",
    "tlf_ia_unique_listings": "Number of unique listings needed for Interim Analysis (IA)",
    "tlf_ia_unique_tables": "Number of unique tables needed for Interim Analysis (IA)",
    "tlf_repeat_figures": "Number of repeat figures needed",
    "tlf_repeat_listings": "Number of repeat listings needed",
    "tlf_repeat_tables": "Number of repeat tables needed",
    "tlf_unique_figures": "Number of unique figures needed",
    "tlf_unique_listings": "Number of unique figures needed",
    "tlf_unique_tables": "Number of unique figures needed",
    "total_dur": "Total duration of all phases of study (months)",
    "num_visits": "Number of visits per subject",
    "avg_unscheduled_vists": "Average number of unscheduled visits per subject",
    "dropout_rate": "Withdrawal rate of enrolled subjects",


}
# Mathematical or business rules used to derive each calculated field.
# Fields without entries here will render with an em dash, indicating that the
# value came directly from the source materials without additional math.
FIELD_FORMULAS = {
    "adam_fr": "subj_dur * 1.5",
    "crf_pages_complete": "num_visits * crf_pages_per_visit",
    "crf_pages_total": "num_complete * (crf_pages_complete + avg_unscheduled_visits * crf_pages_per_visit) + num_withdrawn * crf_pages_withdrawn + num_screen_fail * crf_pages_screen_fail",
    "crf_pages_withdrawn": "crf_pages_complete / 2",
    "dsur_years": "floor(total_dur / 12)",
    "investigator_years": "floor(total_dur / 12)",
    "num_complete": "num_subj * (1 - withdrawal_rate)",
    "num_screen_fail": "num_screened * screen_failure_rate",
    "num_unique_terms_aemh": "num_subj * 10 * 0.05",
    "num_unique_terms_cm": "num_subj * 8 * 0.3",
    "num_withdrawn": "num_subj * dropout_rate",
    "sdtm_fr": "subj_dur * 3",
    "num_dmc_meet": "ceil(subj_dur / 6)",
    "tlf_final_fr": "subj_dur",
    "tlf_dmc_fr": "num_dmc_meet",
    "tlf_dmc_repeat_figures": "floor(tlf_final_repeat_figures * 0.6)",
    "tlf_dmc_repeat_listings": "floor(tlf_final_repeat_listings * 0.6)",
    "tlf_dmc_repeat_tables": "floor(tlf_final_repeat_tables * 0.6)",
    "tlf_dmc_unique_figures": "floor(tlf_final_unique_figures * 0.6)",
    "tlf_dmc_unique_listings": "floor(tlf_final_unique_listings * 0.6)",
    "tlf_dmc_unique_tables": "floor(tlf_final_unique_tables * 0.6)",
    "sdtm_dmc_fr": "num_dmc_meet",
    "adam_dmc_fr": "num_dmc_meet",
    "tlf_ia_repeat_figures": "floor(tlf_final_repeat_figures * 0.75)",
    "tlf_ia_repeat_listings": "floor(tlf_final_repeat_listings * 0.75)",
    "tlf_ia_repeat_tables": "floor(tlf_final_repeat_tables * 0.75)",
    "tlf_ia_unique_figures": "floor(tlf_final_unique_figures * 0.75)",
    "tlf_ia_unique_listings": "floor(tlf_final_unique_listings * 0.75)",
    "tlf_ia_unique_tables": "floor(tlf_final_unique_tables * 0.75)",
}
# Free-form implementation guidance, hints, or suggested values that help users
# understand how to populate each field after extraction.
FIELD_NOTES = {
    "adam_fr": "Benchmark = ~1.5 refreshes/month",
    "crf_pages_complete": "Benchmark = 10 pages per visit",
    "crf_pages_withdrawn": "Assuming half the number of pages for withdrawn subjects",
    "num_unique_terms_aemh": "Assuming 10 AEs per subject with 0.05 unique rate",
    "num_unique_terms_cm": "Assuming 8 CM per subject with 0.3 unique rate",
    "sdtm_fr": "Benchmark = ~3 refreshes/month",
    "tlf_final_fr": "Assuming 1 refresh/month",
    "tlf_dmc_fr": "Assuming 1 refresh/meeting",
    "tlf_dmc_repeat_figures": "Assuming 60% of total TLFs needed for DMC",
    "tlf_dmc_repeat_listings": "Assuming 60% of total TLFs needed for DMC",
    "tlf_dmc_repeat_tables": "Assuming 60% of total TLFs needed for DMC",
    "tlf_dmc_unique_figures": "Assuming 60% of total TLFs needed for DMC",
    "tlf_dmc_unique_listings": "Assuming 60% of total TLFs needed for DMC",
    "tlf_dmc_unique_tables": "Assuming 60% of total TLFs needed for DMC",
    "sdtm_dmc_fr": "Assuming 1 refresh/meeting",
    "adam_dmc_fr": "Assuming 1 refresh/meeting",
    "tlf_ia_repeat_figures": "Assuming 75% of total TLFs needed for IA",
    "tlf_ia_repeat_listings": "Assuming 75% of total TLFs needed for IA",
    "tlf_ia_repeat_tables": "Assuming 75% of total TLFs needed for IA",
    "tlf_ia_unique_figures": "Assuming 75% of total TLFs needed for IA",
    "tlf_ia_unique_listings": "Assuming 75% of total TLFs needed for IA",
    "tlf_ia_unique_tables": "Assuming 75% of total TLFs needed for IA",
}


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "templates_xlsx", "template_full.xlsx")

SHEETS_MAP = {"biostats": ["Study Information", "Biostatistics and Programming"],
              "data_management": ["Study Information", "Clinical Data Management"],
              "project_management": ["Study Information", "Project Management"],
              "conform": ["Study Information", "CONFORM Informatics"],
              }


def allowed_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )

def run_extraction(steps, documents, refresh_opts, dmc_opts):

    data = {}
    print(2)
    if "conform" in steps:
        data.update(get_data_conform(documents))
    if "project_management" in steps:
        data.update(get_data_pm(documents))
    if "data_management" in steps:
        data.update(get_data_dm(documents))
    if "biostats" in steps:
        data.update(get_data_biostats(documents))

    return data

def run_substeps(steps, data, refresh_opts, dmc_opts):
    # if refresh_opts is a tuple (do_refresh, docs[])
    do_refresh, refresh_docs = refresh_opts
    r = True
    if not refresh_docs:
        r = False        
        
    if "biostats" in steps and do_refresh:
        data.update(calculate_refresh(data, refresh_docs, r))

    # same for DMC:
    do_dmc, dmc_docs = dmc_opts
    d = True
    if not dmc_docs:
        d = False
    if "biostats" in steps and do_dmc:
        data.update(calculate_dmc(data, dmc_docs, d))

    return data



@app.route("/", methods=["GET", "POST"])
def select_types():
    if request.method == "POST":
        chosen = request.form.getlist("steps")
        if not chosen:
            return render_template("select_types.html", error="Pick at least one.")
        session["extraction_steps"] = chosen
        session.pop("extracted", None)
        return redirect(url_for("upload_and_extract"))
    return render_template("select_types.html")


@app.route("/upload", methods=["GET", "POST"])
def upload_and_extract():
    steps = session.get("extraction_steps") or []

    if request.method == "POST":
        # 1) Gather any top‑level uploads
        raw_docs = request.files.getlist("docs")
        documents = [
            {
              "file_bytes": f.read(),
              "format": f.filename.rsplit(".", 1)[1].lower(),
              "name":    os.path.splitext(f.filename)[0]
            }
            for f in raw_docs
            if f and allowed_file(f.filename)
        ]

        # 2) Check sub‑step flags and their helper docs
        do_refresh = request.form.get("calculate_refresh") == "yes"
        refresh_docs = [
            {
              "file_bytes": f.read(),
              "format": f.filename.rsplit(".", 1)[1].lower(),
              "name":    os.path.splitext(f.filename)[0]
            }
            for f in request.files.getlist("refresh_docs")
            if f and allowed_file(f.filename)
        ]

        do_dmc = request.form.get("calculate_dmc") == "yes"
        dmc_docs = [
            {
              "file_bytes": f.read(),
              "format": f.filename.rsplit(".", 1)[1].lower(),
              "name":    os.path.splitext(f.filename)[0]
            }
            for f in request.files.getlist("dmc_docs")
            if f and allowed_file(f.filename)
        ]

        # —————————————————————————————
        # A) SAVE‑CHANGES ONLY (no docs, no flags)
        # —————————————————————————————
        if not documents and not do_refresh and not do_dmc:
            # Merge every non‑control form field into session["extracted"]
            data = session.get("extracted", {}).copy()
            for key, val in request.form.items():
                if key not in (
                    "calculate_refresh",
                    "calculate_dmc",
                    "refresh_file_opt_in",
                    "dmc_file_opt_in"
                ):
                    data[key] = val
            session["extracted"] = data

            # Re‑render results table with blanks for any -1
            display = {
              k: ("" if v in (-1, "-1") else v)
              for k, v in data.items()
            }
            return render_template(
                "results.html",
                results=display,
                descriptions=FIELD_DESCRIPTIONS,
                formulas=FIELD_FORMULAS,
                notes=FIELD_NOTES,
            )
        
        if session.get("base_done") and (do_refresh or do_dmc):
            data = session.get("extracted", {}).copy()
            extract = run_substeps(
                steps,
                data,
                (do_refresh, refresh_docs),
                (do_dmc,     dmc_docs),
            )
            session["extracted"] = extract


            display = {k: ("" if v in (-1, "-1") else v) for k, v in data.items()}
            return render_template(
                "results.html",
                results=display,
                descriptions=FIELD_DESCRIPTIONS,
                formulas=FIELD_FORMULAS,
                notes=FIELD_NOTES,
            )

        
        print(8)
        session.pop("base_done", None)
        data = run_extraction(steps, documents, (do_refresh, refresh_docs), (do_dmc, dmc_docs))  
        session["base_done"] = True
        session["extracted"] = data
        print(1)
        display = {k: ("" if v in (-1, "-1") else v) for k, v in data.items()}
        return render_template(
            "results.html",
            results=display,
            descriptions=FIELD_DESCRIPTIONS,
            formulas=FIELD_FORMULAS,
            notes=FIELD_NOTES,
        )

    # GET → show upload form
    return render_template("upload.html")



@app.route("/export", methods=["POST"])
def export():
    steps = session.get("extraction_steps", [])
    data  = session.get("extracted")
    if not steps or not data:
        return redirect(url_for("select_types"))

    # --- sanitize ---
    sanitized = {
        k: ("" if v == -1 or v == "-1" else v)
        for k, v in data.items()
    }
    
    # 1) Fill the master template into a temp file
    tmp_in = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_in.close()
    populate_template(sanitized,TEMPLATE_PATH,tmp_in.name)

    # 2) Load the populated workbook and decide which sheets to keep
    wb = load_workbook(tmp_in.name)
    to_keep = set()
    for step in steps:
        # grab the list of sheet names for this step
        for sheet_name in SHEETS_MAP.get(step, []):
            if sheet_name in wb.sheetnames:
                to_keep.add(sheet_name)

    # 3) Remove any sheet not in the to_keep set
    for sheet in list(wb.sheetnames):
        if sheet not in to_keep:
            wb.remove(wb[sheet])

    # 4) Save the pruned workbook to another temp file and send it
    tmp_out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_out.close()
    wb.save(tmp_out.name)

    resp = send_file(
        tmp_out.name,
        as_attachment=True,
        download_name="budget_proposal.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    @resp.call_on_close
    def cleanup():
        try:
            os.remove(tmp_in.name)
            os.remove(tmp_out.name)
        except OSError:
            pass

    return resp

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
